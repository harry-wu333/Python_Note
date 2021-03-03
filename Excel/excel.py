from openpyxl import load_workbook, Workbook
import time
import sys
import os
# 将表中时间转化为时间戳
def TimeStamp(input_time):
    timeArray = time.strptime(input_time, "%Y-%m-%d %H:%M:%S")
    # print(timeArray)  # 转化为时间数组
    timeStamp = (int(time.mktime(timeArray)))
    # print(timeStamp)  # 转化为时间戳
    return timeStamp

# 获取充电开始时间
def Begin_time(sheet):
    start_time = 0
    start_timestamp = 0
    max_row = sheet.max_row
    for i in range(3, max_row-1):
            pre_cell = sheet.cell(row=i-1, column=2)
            pre_timestamp = TimeStamp(str(pre_cell.value))

            cell = sheet.cell(row=i, column=2)
            cell_timestamp = TimeStamp(str(cell.value))

            level_cell = sheet.cell(row=i, column=3)
            # print("level_cell: %.2f" % level_cell.value + "\n")

            if cell_timestamp-pre_timestamp > 1800 and level_cell.value < 10:
                start_time = cell.value
                start_timestamp = cell_timestamp
                break

    return i, start_time, start_timestamp

# 获取Level 100%充电结束时间
def Level_end_time(sheet, start):
    level_end_time = 0
    level_end_timestamp = 0
    for i in range(start, sheet.max_row-1):
        cell = sheet.cell(row=i, column=3)
        if cell.value >= 100:
            level_end_time = sheet.cell(row=i, column=2).value
            level_end_timestamp = TimeStamp(str(level_end_time))
            break
    return level_end_time, level_end_timestamp

# 获取RawLevel 100%充电结束时间
def Raw_end_time(sheet, start):
    raw_end_time = 0
    raw_end_timestamp = 0
    for i in range(start, sheet.max_row - 1):
        cell = sheet.cell(row=i, column=4)
        if cell.value >= 100:
            raw_end_time = sheet.cell(row=i, column=2).value
            raw_end_timestamp = TimeStamp(str(raw_end_time))
            break
    return raw_end_time, raw_end_timestamp

# 获取GG 100%充电结束时间
def GG_end_time(sheet, start):
    gg_end_time = 0
    gg_end_timestamp = 0
    for i in range(start, sheet.max_row - 1):
        cell = sheet.cell(row=i, column=5)
        if cell.value == 1:
            gg_end_time = sheet.cell(row=i, column=2).value
            gg_end_timestamp = TimeStamp(str(gg_end_time))
            break
    return gg_end_time, gg_end_timestamp

# 计算并输出三种充电时间数据

def Charge_time(start_time, end_time):
    result = 0
    result = (end_time - start_time)/3600
    return result

# 主函数
def main():
    filename = sys.argv[1]
    realpath = os.path.realpath(filename)
    dirpath = os.path.dirname(realpath)
    print("工作表名称为: " + realpath)
    workbook = load_workbook(filename=realpath)# 获取工作表
    # print(workbook.sheetnames)
    sheet = workbook['Sheet1']
    start, start_time, start_timestamp = Begin_time(sheet)
    if start_time == 0:
        print("没有找到开始时间！！！\n")
    else:
        print("充电开始时间为: " + str(start_time) + "\n")
        level_end_time, level_end_timestamp = Level_end_time(sheet, start)
        print("Level 100%的时间为: " + str(level_end_time))
        level_result = Charge_time(start_timestamp, level_end_timestamp)
        print("Level 100% 用时:{:.2f}小时\n".format(level_result))

        raw_end_time, raw_end_timestamp = Raw_end_time(sheet, start)
        print("RawLevel 100%的时间为: " + str(raw_end_time))
        raw_result = Charge_time(start_timestamp, raw_end_timestamp)
        print("RawLevel 100% 用时:{:.2f}小时\n".format(raw_result))

        gg_end_time, gg_end_timestamp = GG_end_time(sheet, start)
        print("GG 100%的时间为: " + str(gg_end_time))
        gg_result = Charge_time(start_timestamp, gg_end_timestamp)
        print("GG 100% 用时:{:.2f}小时\n".format(gg_result))

# 将结果写入新表格
    result_workbook = Workbook()
    result_sheet = result_workbook.active
    data = [str(start_time), round(level_result, 2), round(raw_result, 2), round(gg_result, 2)]
    result_sheet.append(data)
    resultpath = dirpath + '\\result.xlsx'
    result_workbook.save(filename=resultpath)



if __name__ == '__main__':
    main()

